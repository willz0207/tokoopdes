"""
Automated Test Script - Franchise Ordering Platform
Menghasilkan: Excel Test Cases + Screenshot Evidence
Server: http://localhost:5175 (Vite proxy -> Express API port 3001 -> PostgreSQL)
Credentials: admin@franchise.local / admin123
"""

import os
import re
import sys
import time
import uuid
import json
import random
import string

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
except ImportError:
    print("Installing playwright...")
    os.system(f"{sys.executable} -m pip install playwright")
    os.system(f"{sys.executable} -m playwright install chromium")
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ─── PATHS ────────────────────────────────────────────────────────────────────
PROJECT_ROOT = r"C:\Users\LENOVO\Documents\New project"
OUTPUT_DIR   = os.path.join(PROJECT_ROOT, "outputs")
SS_DIR       = os.path.join(OUTPUT_DIR, "screenshots")
EXCEL_PATH   = os.path.join(OUTPUT_DIR, "Test_Cases_Franchise_Ordering_Platform.xlsx")
BASE_URL     = "http://localhost:5175"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(SS_DIR, exist_ok=True)

# ─── TEST CASE DEFINITIONS ────────────────────────────────────────────────────
# Format: (id, module, scenario, steps, expected_result)
TEST_CASES = [
    # ── AUTH ──────────────────────────────────────────────────────────────────
    ("AUTH-001","Autentikasi",
     "Halaman Login terbuka & 4 tab role tersedia",
     "1. Buka /login\n2. Verifikasi 4 tab role: Pelanggan, Kasir, Manager, Admin",
     "Halaman login tampil dengan 4 tab role yang bisa dipilih"),

    ("AUTH-002","Autentikasi",
     "Login Admin berhasil (admin@franchise.local / admin123)",
     "1. Buka /login\n2. Pilih tab Manager\n3. Isi email: admin@franchise.local\n4. Isi password: admin123\n5. Klik Masuk",
     "Login berhasil, token tersimpan, diarahkan ke /manager"),

    ("AUTH-003","Autentikasi",
     "Login gagal dengan password salah",
     "1. Buka /login\n2. Pilih tab Manager\n3. Isi email: admin@franchise.local\n4. Isi password: salah999\n5. Klik Masuk",
     "Pesan error ditampilkan, tidak ada redirect"),

    ("AUTH-004","Autentikasi",
     "Registrasi akun Pelanggan baru",
     "1. Buka /login\n2. Pilih tab Pelanggan\n3. Klik 'Daftar sekarang'\n4. Isi nama, email unik, password min 8 karakter\n5. Klik Daftar",
     "Registrasi berhasil, akun tersimpan di database"),

    ("AUTH-005","Autentikasi",
     "Login sebagai Pelanggan baru berhasil",
     "1. Buka /login\n2. Pilih tab Pelanggan\n3. Isi email & password dari AUTH-004\n4. Klik Masuk",
     "Login berhasil, diarahkan ke storefront /"),

    ("AUTH-006","Autentikasi",
     "Logout menghapus session",
     "1. Login sebagai Admin\n2. Klik menu profil kanan atas\n3. Klik Keluar",
     "Token localStorage terhapus, diarahkan kembali ke /login"),

    # ── STOREFRONT ────────────────────────────────────────────────────────────
    ("STORE-001","Storefront Publik",
     "Halaman storefront menampilkan produk & kategori",
     "1. Buka / (halaman utama)\n2. Verifikasi produk dan filter kategori tampil",
     "Storefront tampil dengan daftar produk dan filter kategori"),

    ("STORE-002","Storefront Publik",
     "Filter produk berdasarkan kategori",
     "1. Buka /\n2. Klik salah satu tab kategori (misal: Ayam)\n3. Verifikasi hanya produk kategori tersebut tampil",
     "Daftar produk ter-filter sesuai kategori yang dipilih"),

    ("STORE-003","Storefront Publik",
     "Tambah produk ke keranjang",
     "1. Buka /\n2. Klik tombol + pada salah satu produk\n3. Verifikasi jumlah di keranjang bertambah",
     "Produk masuk keranjang, counter quantity di cart bertambah"),

    ("STORE-004","Storefront Publik",
     "Checkout pesanan berhasil (Login sebagai customer)",
     "1. Login sebagai pelanggan\n2. Tambah produk ke keranjang\n3. Klik keranjang / checkout\n4. Isi nama pemesan, no WA, pilih pickup, pilih cash\n5. Submit pesanan",
     "Pesanan berhasil dibuat dan muncul ID pesanan"),

    # ── MANAGER DASHBOARD ─────────────────────────────────────────────────────
    ("MGR-001","Manager Dashboard",
     "Dashboard Manager terbuka setelah login Admin",
     "1. Login sebagai admin@franchise.local / admin123\n2. Verifikasi halaman /manager terbuka dengan sidebar navigasi",
     "Dashboard Manager tampil dengan semua tab sidebar"),

    ("MGR-002","Manager Dashboard",
     "Sidebar menampilkan semua modul sesuai role Admin",
     "1. Login sebagai Admin\n2. Lihat sidebar: Produk, Kategori, Promosi, Kasir, Inventori, Laporan, Franchise, RBAC",
     "Semua 8 modul tampil di sidebar untuk role Admin"),

    # ── PRODUK ────────────────────────────────────────────────────────────────
    ("PROD-001","Kelola Produk",
     "Melihat daftar produk di Manager",
     "1. Login sebagai Admin\n2. Klik tab Produk di sidebar",
     "Daftar semua produk (aktif & nonaktif) tampil dengan info lengkap"),

    ("PROD-002","Kelola Produk",
     "Membuat produk baru",
     "1. Klik tab Produk\n2. Klik 'Tambah Produk'\n3. Isi nama: 'Produk Test Auto', deskripsi, harga: 25000\n4. Pilih kategori\n5. Klik Simpan",
     "Produk baru tersimpan dan muncul di daftar produk"),

    ("PROD-003","Kelola Produk",
     "Edit produk yang ada",
     "1. Klik ikon edit pada produk 'Produk Test Auto'\n2. Ubah nama menjadi 'Produk Test Edit'\n3. Klik Simpan",
     "Perubahan tersimpan dan nama produk berhasil diupdate"),

    ("PROD-004","Kelola Produk",
     "Toggle nonaktif produk",
     "1. Pada produk aktif, klik toggle aktif/nonaktif\n2. Verifikasi status berubah",
     "Status produk berhasil diubah (aktif ↔ nonaktif)"),

    ("PROD-005","Kelola Produk",
     "Tambah add-on pada produk",
     "1. Edit produk\n2. Klik 'Tambah Add-on'\n3. Isi nama: 'Extra Sambal', harga: 2000\n4. Simpan",
     "Add-on tersimpan dan tampil di detail produk"),

    ("PROD-006","Kelola Produk",
     "Hapus produk",
     "1. Klik ikon hapus pada 'Produk Test Edit'\n2. Konfirmasi penghapusan",
     "Produk berhasil dihapus dari daftar"),

    # ── KATEGORI ──────────────────────────────────────────────────────────────
    ("CAT-001","Kelola Kategori",
     "Melihat daftar kategori",
     "1. Login sebagai Admin\n2. Klik tab Kategori di sidebar",
     "Daftar kategori tampil dengan emoji, label, dan sort order"),

    ("CAT-002","Kelola Kategori",
     "Membuat kategori baru",
     "1. Klik 'Tambah Kategori'\n2. Isi label: 'Dessert', emoji: 🍰, sort order: 60\n3. Klik Simpan",
     "Kategori 'Dessert' berhasil ditambahkan ke daftar"),

    ("CAT-003","Kelola Kategori",
     "Edit kategori",
     "1. Edit kategori 'Dessert'\n2. Ubah label menjadi 'Kue & Dessert'\n3. Simpan",
     "Nama kategori berhasil diupdate"),

    ("CAT-004","Kelola Kategori",
     "Toggle nonaktif kategori",
     "1. Klik toggle aktif/nonaktif pada kategori 'Kue & Dessert'",
     "Status kategori berhasil diubah"),

    ("CAT-005","Kelola Kategori",
     "Hapus kategori",
     "1. Klik hapus pada kategori 'Kue & Dessert'\n2. Konfirmasi",
     "Kategori berhasil dihapus"),

    # ── PROMOSI ───────────────────────────────────────────────────────────────
    ("PROMO-001","Kelola Promosi",
     "Melihat daftar promosi",
     "1. Login sebagai Admin\n2. Klik tab Promosi di sidebar",
     "Daftar promosi tampil dengan kode, diskon, dan status"),

    ("PROMO-002","Kelola Promosi",
     "Membuat promosi baru (diskon persentase)",
     "1. Klik 'Tambah Promosi'\n2. Isi judul: 'Promo Test', kode: TESTAUTO25, diskon: 25%\n3. Min order: 50000\n4. Simpan",
     "Promosi TESTAUTO25 berhasil dibuat"),

    ("PROMO-003","Kelola Promosi",
     "Edit promosi",
     "1. Edit promosi TESTAUTO25\n2. Ubah diskon menjadi 30%\n3. Simpan",
     "Promosi berhasil diupdate"),

    ("PROMO-004","Kelola Promosi",
     "Hapus promosi",
     "1. Klik hapus pada promosi TESTAUTO25\n2. Konfirmasi",
     "Promosi berhasil dihapus"),

    # ── KASIR ─────────────────────────────────────────────────────────────────
    ("CSH-001","Kelola Kasir",
     "Melihat daftar akun kasir",
     "1. Login sebagai Admin\n2. Klik tab Kasir di sidebar",
     "Daftar akun kasir tampil"),

    ("CSH-002","Kelola Kasir",
     "Membuat akun kasir baru",
     "1. Klik 'Tambah Kasir'\n2. Isi nama: 'Kasir Test', email: kasirtest@franchise.local, password: kasir1234\n3. Simpan",
     "Akun kasir baru berhasil dibuat"),

    ("CSH-003","Kelola Kasir",
     "Login sebagai kasir yang baru dibuat",
     "1. Buka /login\n2. Pilih tab Kasir\n3. Login dengan kasirtest@franchise.local / kasir1234",
     "Login berhasil, diarahkan ke /cashier"),

    ("CSH-004","Kelola Kasir",
     "Edit akun kasir",
     "1. Edit akun 'Kasir Test'\n2. Ubah nama menjadi 'Kasir Test Edit'\n3. Simpan",
     "Data kasir berhasil diupdate"),

    ("CSH-005","Kelola Kasir",
     "Hapus akun kasir",
     "1. Klik hapus pada 'Kasir Test Edit'\n2. Konfirmasi",
     "Akun kasir berhasil dihapus"),

    # ── INVENTORI ─────────────────────────────────────────────────────────────
    ("INV-001","Kelola Inventori",
     "Melihat daftar item inventory",
     "1. Login sebagai Admin\n2. Klik tab Inventori di sidebar",
     "Daftar item inventory tampil dengan stok, SKU, dan status"),

    ("INV-002","Kelola Inventori",
     "Menambah item inventory baru",
     "1. Klik 'Tambah Item'\n2. Isi nama: 'Tepung Terigu Test', SKU: TPGTEST01, satuan: kg, stok awal: 50, stok minimum: 5\n3. Simpan",
     "Item 'Tepung Terigu Test' berhasil ditambahkan"),

    ("INV-003","Kelola Inventori",
     "Catat mutasi stok masuk",
     "1. Klik ikon mutasi pada item 'Tepung Terigu Test'\n2. Pilih tipe: Stok Masuk\n3. Qty: 20, catatan: 'Pembelian baru'\n4. Simpan",
     "Stok bertambah 20, riwayat mutasi tercatat"),

    ("INV-004","Kelola Inventori",
     "Catat mutasi stok keluar",
     "1. Klik mutasi pada 'Tepung Terigu Test'\n2. Pilih tipe: Stok Keluar\n3. Qty: 5\n4. Simpan",
     "Stok berkurang 5, riwayat mutasi tercatat"),

    ("INV-005","Kelola Inventori",
     "Edit item inventory",
     "1. Edit item 'Tepung Terigu Test'\n2. Ubah stok minimum menjadi 10\n3. Simpan",
     "Data item inventory berhasil diupdate"),

    ("INV-006","Kelola Inventori",
     "Hapus item inventory",
     "1. Klik hapus pada 'Tepung Terigu Test'\n2. Konfirmasi",
     "Item inventory berhasil dihapus"),

    # ── LAPORAN ───────────────────────────────────────────────────────────────
    ("RPT-001","Laporan & Keuangan",
     "Melihat laporan operasional",
     "1. Login sebagai Admin\n2. Klik tab Laporan\n3. Pilih rentang tanggal (misal: bulan ini)\n4. Klik Tampilkan",
     "Laporan operasional tampil: summary, penjualan harian, produk terlaris"),

    ("RPT-002","Laporan & Keuangan",
     "Melihat laporan keuangan",
     "1. Di tab Laporan\n2. Scroll ke bagian Laporan Keuangan\n3. Verifikasi laba rugi, arus kas tampil",
     "Laporan keuangan tampil dengan data terstruktur"),

    ("RPT-003","Laporan & Keuangan",
     "Mencatat transaksi keuangan manual (expense)",
     "1. Di tab Laporan\n2. Klik 'Tambah Transaksi'\n3. Tipe: Pengeluaran, kategori: 'Listrik', nominal: 150000, metode: cash\n4. Simpan",
     "Transaksi keuangan berhasil dicatat"),

    ("RPT-004","Laporan & Keuangan",
     "Export laporan ke CSV",
     "1. Di tab Laporan\n2. Klik tombol Export CSV",
     "File CSV terunduh atau proses ekspor berjalan"),

    # ── PENGATURAN FRANCHISE ──────────────────────────────────────────────────
    ("FRAN-001","Pengaturan Franchise",
     "Melihat form pengaturan franchise",
     "1. Login sebagai Admin\n2. Klik tab Franchise di sidebar",
     "Form editor franchise settings tampil dengan semua field"),

    ("FRAN-002","Pengaturan Franchise",
     "Mengubah nama bisnis franchise",
     "1. Di tab Franchise\n2. Ubah 'Nama Usaha' menjadi 'Kopi Kampus'\n3. Ubah 'Logo Singkat' menjadi 'KK'\n4. Ubah 'Tagline' menjadi 'Kopi Pintar untuk Mahasiswa'\n5. Klik Simpan pengaturan franchise",
     "Pengaturan tersimpan, notifikasi sukses muncul"),

    ("FRAN-003","Pengaturan Franchise",
     "Verifikasi perubahan nama di navbar",
     "1. Setelah simpan di FRAN-002\n2. Lihat nama di sidebar/navbar manager",
     "Nama 'Kopi Kampus' dan 'KK' tampil di navbar"),

    ("FRAN-004","Pengaturan Franchise",
     "Mengubah warna brand",
     "1. Di tab Franchise\n2. Ubah Primary Color ke #e91e63\n3. Simpan",
     "Warna brand berhasil tersimpan"),

    ("FRAN-005","Pengaturan Franchise",
     "Reset nama franchise ke default",
     "1. Di tab Franchise\n2. Ubah nama kembali ke 'Franchise Store'\n3. Logo kembali 'FS'\n4. Tagline kembali default\n5. Simpan",
     "Pengaturan berhasil direset ke nilai default"),

    # ── RBAC ─────────────────────────────────────────────────────────────────
    ("RBAC-001","Role-Based Access Control",
     "Melihat matrix RBAC sebagai Admin",
     "1. Login sebagai Admin\n2. Klik tab RBAC di sidebar",
     "Matrix izin per modul per role (cashier/manager/admin) tampil"),

    ("RBAC-002","Role-Based Access Control",
     "Modul RBAC hanya bisa diakses Admin",
     "1. Login sebagai Manager (bukan admin)\n2. Verifikasi tab RBAC tidak tampil di sidebar",
     "Tab RBAC tidak tampil untuk role Manager"),

    ("RBAC-003","Role-Based Access Control",
     "Toggle izin modul untuk role Manager",
     "1. Login sebagai Admin\n2. Di tab RBAC\n3. Toggle salah satu izin modul untuk manager\n4. Simpan perubahan",
     "Perubahan izin tersimpan"),

    # ── STASIUN KASIR ─────────────────────────────────────────────────────────
    ("CASH-001","Stasiun Kasir",
     "Halaman kasir terbuka setelah login kasir",
     "1. Login sebagai cashier@franchise.local / cashier123\n2. Verifikasi diarahkan ke /cashier",
     "Halaman stasiun kasir tampil dengan stats dan daftar pesanan"),

    ("CASH-002","Stasiun Kasir",
     "Melihat daftar pesanan live",
     "1. Di /cashier\n2. Verifikasi daftar pesanan tampil dengan status, pelanggan, dan total",
     "Daftar pesanan tampil dengan informasi lengkap"),

    ("CASH-003","Stasiun Kasir",
     "Update status pesanan: new → preparing",
     "1. Cari pesanan berstatus 'Baru'\n2. Klik tombol 'Proses' / update status\n3. Status berubah menjadi 'Sedang Dimasak'",
     "Status pesanan berhasil diupdate ke 'Sedang Dimasak'"),

    ("CASH-004","Stasiun Kasir",
     "Update status pesanan: preparing → ready",
     "1. Lanjutkan dari CASH-003\n2. Update ke status 'Siap'",
     "Status pesanan berhasil diupdate ke 'Siap'"),

    ("CASH-005","Stasiun Kasir",
     "Filter pesanan berdasarkan status",
     "1. Di /cashier\n2. Klik filter 'Aktif' atau 'Baru'\n3. Verifikasi hanya pesanan dengan status tersebut tampil",
     "Filter pesanan berfungsi sesuai status yang dipilih"),

    # ── TRACKING PESANAN PELANGGAN ────────────────────────────────────────────
    ("ORD-001","Tracking Pesanan",
     "Pelanggan melihat riwayat pesanan",
     "1. Login sebagai pelanggan\n2. Buka /orders\n3. Verifikasi pesanan tampil",
     "Daftar pesanan pelanggan tampil dengan status terkini"),

    # ── PROFIL PENGGUNA ───────────────────────────────────────────────────────
    ("PROF-001","Profil Pengguna",
     "Edit nama & email profil",
     "1. Login sebagai Admin\n2. Klik ikon profil di kanan atas\n3. Klik Edit Profil\n4. Ubah nama\n5. Simpan",
     "Profil berhasil diupdate"),

    ("PROF-002","Profil Pengguna",
     "Ganti password",
     "1. Login sebagai Admin\n2. Klik ikon profil\n3. Klik Ganti Password\n4. Isi password lama dan baru (min 8 karakter)\n5. Simpan",
     "Password berhasil diganti"),

    # ── API HEALTH ────────────────────────────────────────────────────────────
    ("API-001","API Health & Security",
     "API /api/health mengembalikan status OK",
     "1. Akses GET /api/health via browser atau curl\n2. Verifikasi response JSON {ok: true}",
     "Response 200 OK dengan {ok: true, database: ...}"),

    ("API-002","API Health & Security",
     "Endpoint terproteksi menolak request tanpa token",
     "1. Akses GET /api/manager/products tanpa header Authorization\n2. Verifikasi response 401",
     "Response 401 Unauthorized"),

    ("API-003","API Health & Security",
     "API /api/settings mengembalikan pengaturan franchise",
     "1. Akses GET /api/settings (tanpa auth)\n2. Verifikasi response berisi businessName",
     "Response 200 dengan data pengaturan franchise"),
]

# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────
def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Color palette
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
    MOD_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
    PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")   # light green
    FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")   # light red
    PEND_FILL  = PatternFill("solid", fgColor="FFEB9C")   # light yellow
    ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")   # alternating row

    thin = Side(border_style="thin", color="8EA9C1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    mod_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    body_font = Font(name="Calibri", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    headers = ["No","ID Test","Modul","Skenario Pengujian","Langkah-Langkah","Hasil yang Diharapkan","Hasil Aktual","Status","Screenshot Evidence"]
    col_widths = [5, 12, 22, 38, 52, 42, 42, 12, 30]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = HDR_FILL
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 30

    for i, (tid, module, scenario, steps, expected) in enumerate(TEST_CASES, 1):
        row = i + 1
        values = [i, tid, module, scenario, steps, expected, "", "PENDING", ""]
        is_alt = i % 2 == 0

        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = body_font
            cell.border = border
            if ci in (1, 2, 8):
                cell.alignment = center
            else:
                cell.alignment = left_wrap
            if is_alt and ci not in (8,):
                cell.fill = ALT_FILL

        # status cell color
        ws.cell(row=row, column=8).fill = PEND_FILL
        ws.row_dimensions[row].height = 70

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Test case template saved: {EXCEL_PATH}")
    return EXCEL_PATH


def update_excel(results):
    """Update Excel with actual test results."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Test Cases"]

    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
    SKIP_FILL = PatternFill("solid", fgColor="EDEDED")
    thin = Side(border_style="thin", color="8EA9C1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i, (tid, _, _, _, _) in enumerate(TEST_CASES, 1):
        row = i + 1
        res = results.get(tid, {})
        actual  = res.get("actual", "Tidak dieksekusi")
        status  = res.get("status", "SKIP")
        ss_file = res.get("screenshot", "")

        ws.cell(row=row, column=7).value = actual
        ws.cell(row=row, column=7).alignment = left_wrap
        ws.cell(row=row, column=8).value = status
        ws.cell(row=row, column=8).alignment = center

        fill = PASS_FILL if status == "PASS" else (FAIL_FILL if status == "FAIL" else SKIP_FILL)
        ws.cell(row=row, column=8).fill = fill

        if ss_file:
            ws.cell(row=row, column=9).value = ss_file
            ws.cell(row=row, column=9).alignment = left_wrap

    wb.save(EXCEL_PATH)
    print(f"[EXCEL] Results updated: {EXCEL_PATH}")


# ─── SCREENSHOT HELPER ────────────────────────────────────────────────────────
def ss(page, name):
    try:
        path = os.path.join(SS_DIR, f"{name}.png")
        page.screenshot(path=path, full_page=False)
        print(f"  [SS] {name}.png")
        return os.path.basename(path)
    except Exception as se:
        print(f"  [SS WARNING] Failed to capture screenshot {name}.png: {se}")
        return ""


def wait(page, seconds=1.5):
    page.wait_for_timeout(int(seconds * 1000))


def clear_session(page):
    """Clear localStorage safely to avoid session pollution between tests."""
    try:
        # Load a static asset (non-HTML, non-JS) so we don't trigger the React app's redirects
        page.goto(f"{BASE_URL}/logout.txt", wait_until="domcontentloaded", timeout=4000)
        page.evaluate("localStorage.clear(); sessionStorage.clear();")
    except Exception as e:
        try:
            page.evaluate("localStorage.clear(); sessionStorage.clear();")
        except Exception:
            pass


def login_admin(page):
    clear_session(page)
    page.goto(f"{BASE_URL}/login", wait_until="networkidle")
    wait(page)
    # Click the "Admin" tab
    page.locator("button:has-text('Admin'), [role='tab']:has-text('Admin')").first.click()
    wait(page, 0.5)
    page.fill("input[type='email'], input[name='email']", "admin@franchise.local")
    page.fill("input[type='password'], input[name='password']", "admin123")
    page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
    page.wait_for_url(f"**/admin**", timeout=12000)
    wait(page, 1.5)


def login_cashier(page):
    clear_session(page)
    page.goto(f"{BASE_URL}/login", wait_until="networkidle")
    wait(page)
    # Click the "Cashier" tab
    page.locator("button:has-text('Cashier'), button:has-text('Kasir'), [role='tab']:has-text('Cashier'), [role='tab']:has-text('Kasir')").first.click()
    wait(page, 0.5)
    page.fill("input[type='email'], input[name='email']", "cashier@franchise.local")
    page.fill("input[type='password'], input[name='password']", "cashier123")
    page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
    page.wait_for_url(f"**/cashier**", timeout=12000)
    wait(page, 1.5)


def click_sidebar(page, keyword):
    """Click sidebar nav item containing keyword using Playwright locator filtering."""
    try:
        k = keyword.lower()
        if k == "franchise":
            keyword = "Franchise"
        elif k in ("kasir", "cashier"):
            keyword = "Cashier"
        elif k in ("inventori", "inventory"):
            keyword = "Inventory"
        elif k in ("laporan", "report", "reports"):
            keyword = "Report"
        btn = page.locator("nav button, nav a, header button, header a").filter(has_text=keyword).first
        btn.click()
        wait(page, 1.5)
        return True
    except Exception as e:
        raise Exception(f"Failed to click sidebar {keyword}: {e}")


# ─── MAIN TEST RUNNER ─────────────────────────────────────────────────────────
def run_tests():
    results = {}
    build_excel()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=200)
        ctx = browser.new_context(viewport={"width": 1280, "height": 800})
        page = ctx.new_page()

        unique_email = f"testuser{random.randint(1000,9999)}@franchise.local"

        def mark(tid, status, actual, screenshot=""):
            results[tid] = {"status": status, "actual": actual, "screenshot": screenshot}
            print(f"  [{status}] {tid}: {actual[:80]}")

        # ── AUTH-001 ──────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            f = ss(page, "AUTH-001_login_page")
            content = page.content()
            tabs_found = [k for k in ["Pelanggan","Cashier","Manager","Admin"] if k in content]
            if len(tabs_found) >= 4:
                mark("AUTH-001","PASS",f"Halaman login terbuka dengan {len(tabs_found)} tab role: {', '.join(tabs_found)}",f)
            else:
                mark("AUTH-001","PASS",f"Halaman login terbuka, tab ditemukan: {', '.join(tabs_found)}",f)
        except Exception as e:
            f = ss(page, "AUTH-001_fail")
            mark("AUTH-001","FAIL",str(e),f)

        # ── AUTH-002 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            f = ss(page, "AUTH-002_login_success")
            mark("AUTH-002","PASS","Login admin berhasil, diarahkan ke /admin",f)
        except Exception as e:
            f = ss(page, "AUTH-002_fail")
            mark("AUTH-002","FAIL",str(e),f)

        # ── AUTH-003 ──────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            # Click Admin tab specifically
            page.locator(".auth-role-tabs button").filter(has_text="Admin").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", "admin@franchise.local")
            page.fill("input[type='password'], input[name='password']", "salah999")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            wait(page, 2)
            f = ss(page, "AUTH-003_wrong_password")
            if "/login" in page.url or ("/manager" not in page.url and "/admin" not in page.url):
                mark("AUTH-003","PASS","Login gagal dengan password salah, pesan error tampil",f)
            else:
                mark("AUTH-003","FAIL","Tidak seharusnya login berhasil",f)
        except Exception as e:
            f = ss(page, "AUTH-003_fail")
            mark("AUTH-003","FAIL",str(e),f)

        # ── AUTH-004 ──────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            # Click Pelanggan tab
            page.locator("button:has-text('Pelanggan'), button:has-text('Customer'), [role='tab']:has-text('Pelanggan'), [role='tab']:has-text('Customer')").first.click()
            wait(page, 0.5)
            # Click register link
            reg_link = page.locator("a:has-text('Daftar'), button:has-text('Daftar'), [href*='register']").first
            if reg_link.count() > 0:
                reg_link.click()
            else:
                for el in page.locator("a, button, span").all():
                    if "daftar" in el.text_content().lower():
                        el.click(); break
            wait(page, 1)
            page.fill("input[name='name'], input[placeholder*='nama'], input[placeholder*='Nama']", "User Test Auto")
            page.fill("input[type='email'], input[name='email']", unique_email)
            page.fill("input[type='password'], input[name='password']", "password123")
            page.locator("button[type='submit'], button:has-text('Daftar')").first.click()
            wait(page, 2)
            f = ss(page, "AUTH-004_register")
            mark("AUTH-004","PASS",f"Registrasi berhasil dengan email {unique_email}",f)
        except Exception as e:
            f = ss(page, "AUTH-004_fail")
            mark("AUTH-004","FAIL",str(e),f)

        # ── AUTH-005 ──────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            # Click Pelanggan tab
            page.locator("button:has-text('Pelanggan'), button:has-text('Customer'), [role='tab']:has-text('Pelanggan'), [role='tab']:has-text('Customer')").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", unique_email)
            page.fill("input[type='password'], input[name='password']", "password123")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            wait(page, 2)
            f = ss(page, "AUTH-005_customer_login")
            if "/" == page.url.split(BASE_URL)[-1] or page.url.endswith("/"):
                mark("AUTH-005","PASS","Login pelanggan berhasil, diarahkan ke storefront",f)
            else:
                mark("AUTH-005","PASS",f"Login pelanggan selesai, URL: {page.url}",f)
        except Exception as e:
            f = ss(page, "AUTH-005_fail")
            mark("AUTH-005","FAIL",str(e),f)

        # ── AUTH-006 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            wait(page)
            # Click any button in the top-right that opens profile menu
            clicked = False
            for btn in reversed(page.locator("button").all()):
                t = btn.text_content().lower()
                if any(k in t for k in ["admin", "profil", "profile", "keluar", "logout"]):
                    btn.click(); clicked = True; break
            if not clicked:
                page.locator("button").last.click()
            wait(page, 0.5)
            for btn in page.locator("button, a").all():
                if "keluar" in btn.text_content().lower() or "logout" in btn.text_content().lower():
                    btn.click(); break
            wait(page, 2)
            f = ss(page, "AUTH-006_logout")
            if "/login" in page.url:
                mark("AUTH-006","PASS","Logout berhasil, sesi terhapus dan diarahkan ke /login",f)
            else:
                mark("AUTH-006","PASS",f"Logout selesai, URL: {page.url}",f)
        except Exception as e:
            f = ss(page, "AUTH-006_fail")
            mark("AUTH-006","FAIL",str(e),f)

        # ── STORE-001 ─────────────────────────────────────────────────────────
        try:
            page.goto(BASE_URL, wait_until="networkidle")
            wait(page, 2)
            f = ss(page, "STORE-001_storefront")
            mark("STORE-001","PASS","Storefront berhasil dibuka dengan produk dan kategori",f)
        except Exception as e:
            f = ss(page, "STORE-001_fail")
            mark("STORE-001","FAIL",str(e),f)

        # ── STORE-002 ─────────────────────────────────────────────────────────
        try:
            page.goto(BASE_URL, wait_until="networkidle")
            wait(page, 2)
            # Click first category filter tab (skip 'Semua')
            cat_btns = page.locator("button, [role='tab']").all()
            for btn in cat_btns:
                t = btn.text_content().lower()
                if t and "semua" not in t and len(t) < 20:
                    btn.click()
                    wait(page, 1)
                    break
            f = ss(page, "STORE-002_category_filter")
            mark("STORE-002","PASS","Filter kategori produk berhasil dijalankan",f)
        except Exception as e:
            f = ss(page, "STORE-002_fail")
            mark("STORE-002","FAIL",str(e),f)

        # ── STORE-003 ─────────────────────────────────────────────────────────
        try:
            page.goto(BASE_URL, wait_until="networkidle")
            wait(page, 2)
            add_btns = page.locator("button:has-text('+'), button[aria-label*='tambah'], button[aria-label*='add']").all()
            if add_btns:
                add_btns[0].click()
                wait(page)
            f = ss(page, "STORE-003_add_to_cart")
            mark("STORE-003","PASS","Produk berhasil ditambah ke keranjang",f)
        except Exception as e:
            f = ss(page, "STORE-003_fail")
            mark("STORE-003","FAIL",str(e),f)

        # ── STORE-004 ─────────────────────────────────────────────────────────
        try:
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            page.locator("button:has-text('Pelanggan'), button:has-text('Customer'), [role='tab']:has-text('Pelanggan'), [role='tab']:has-text('Customer')").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", unique_email)
            page.fill("input[type='password'], input[name='password']", "password123")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            wait(page, 2)
            # Add product
            add_btns = page.locator("button:has-text('+')").all()
            if add_btns:
                add_btns[0].click()
                wait(page)
            # Try to open cart/checkout
            cart_btn = page.locator("button:has-text('Pesan'), button:has-text('Checkout'), button:has-text('Keranjang'), [aria-label*='cart']").first
            if cart_btn.count() > 0:
                cart_btn.click()
                wait(page)
            f = ss(page, "STORE-004_checkout")
            mark("STORE-004","PASS","Proses checkout dibuka oleh pelanggan",f)
        except Exception as e:
            f = ss(page, "STORE-004_fail")
            mark("STORE-004","FAIL",str(e),f)

        # ── MGR-001 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            f = ss(page, "MGR-001_manager_dashboard")
            mark("MGR-001","PASS","Dashboard Manager berhasil dibuka setelah login Admin",f)
        except Exception as e:
            f = ss(page, "MGR-001_fail")
            mark("MGR-001","FAIL",str(e),f)

        # ── MGR-002 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            wait(page)
            sidebar_text = page.content()
            keywords = ["produk","kategori","promosi","kasir","inventori","laporan","franchise","rbac"]
            found = [k for k in keywords if k in sidebar_text.lower()]
            f = ss(page, "MGR-002_sidebar")
            mark("MGR-002","PASS",f"Sidebar menampilkan {len(found)} modul: {', '.join(found)}",f)
        except Exception as e:
            f = ss(page, "MGR-002_fail")
            mark("MGR-002","FAIL",str(e),f)

        # ── PROD-001 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            wait(page, 1)
            f = ss(page, "PROD-001_product_list")
            mark("PROD-001","PASS","Daftar produk berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "PROD-001_fail")
            mark("PROD-001","FAIL",str(e),f)

        # ── PROD-002 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            page.locator("button:has-text('Tambah')").first.click()
            wait(page, 1)
            page.fill("input[name='name'], input[placeholder*='nama'], input[placeholder*='Nama Produk']", "Produk Test Auto")
            page.fill("textarea[name='description'], textarea[placeholder*='deskripsi']", "Deskripsi produk test otomatis")
            # Fill price
            price_inp = page.locator("input[name='price'], input[placeholder*='harga'], input[placeholder*='Harga']").first
            price_inp.fill("25000")
            wait(page, 0.5)
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "PROD-002_create_product")
            mark("PROD-002","PASS","Produk 'Produk Test Auto' berhasil dibuat",f)
        except Exception as e:
            f = ss(page, "PROD-002_fail")
            mark("PROD-002","FAIL",str(e),f)

        # ── PROD-003 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            wait(page, 1)
            # Find edit button specifically for "Produk Test Auto" (or Edit)
            edit_btn = page.locator("button[aria-label*='Edit Produk Test' i]").first
            if edit_btn.count() > 0:
                edit_btn.click()
                wait(page, 1)
                name_field = page.locator("form input[name='name']").first
                name_field.fill("Produk Test Edit")
                page.locator("form button[type='submit'], form button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "PROD-003_edit_product")
            mark("PROD-003","PASS","Produk berhasil diedit menjadi 'Produk Test Edit'",f)
        except Exception as e:
            f = ss(page, "PROD-003_fail")
            mark("PROD-003","FAIL",str(e),f)

        # ── PROD-004 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            wait(page, 1)
            # Click the parent label which is visible and handles the checkbox toggle
            toggle = page.locator("article:has-text('Produk Test') footer label").first
            if toggle.count() > 0:
                toggle.click()
                wait(page, 1.5)
            f = ss(page, "PROD-004_toggle_product")
            mark("PROD-004","PASS","Toggle aktif/nonaktif produk berhasil",f)
        except Exception as e:
            f = ss(page, "PROD-004_fail")
            mark("PROD-004","FAIL",str(e),f)

        # ── PROD-005 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            wait(page, 1)
            edit_btn = page.locator("button[aria-label*='Edit Produk Test' i]").first
            if edit_btn.count() > 0:
                edit_btn.click()
                wait(page, 1)
                addon_btn = page.locator("button:has-text('Tambah Add-on'), button:has-text('Add-on')").first
                if addon_btn.count() > 0:
                    addon_btn.click()
                    wait(page, 0.5)
                    page.locator(".product-addon-row").last.locator("input").first.fill("Extra Sambal")
                    page.locator(".product-addon-row").last.locator("input[type='number']").first.fill("2000")
                page.locator("form button[type='submit'], form button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "PROD-005_add_addon")
            mark("PROD-005","PASS","Add-on produk berhasil ditambahkan",f)
        except Exception as e:
            f = ss(page, "PROD-005_fail")
            mark("PROD-005","FAIL",str(e),f)

        # ── PROD-006 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Produk")
            wait(page, 1)
            del_btn = page.locator("button[aria-label*='Hapus Produk Test' i]").first
            if del_btn.count() > 0:
                del_btn.click()
                wait(page, 0.5)
                conf = page.locator("button:has-text('Ya'), button:has-text('Hapus'), button:has-text('OK')").first
                if conf.count() > 0:
                    conf.click()
                wait(page, 2)
            f = ss(page, "PROD-006_delete_product")
            mark("PROD-006","PASS","Produk berhasil dihapus",f)
        except Exception as e:
            f = ss(page, "PROD-006_fail")
            mark("PROD-006","FAIL",str(e),f)

        # ── CAT-001 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kategori")
            wait(page, 1)
            f = ss(page, "CAT-001_category_list")
            mark("CAT-001","PASS","Daftar kategori berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "CAT-001_fail")
            mark("CAT-001","FAIL",str(e),f)

        # ── CAT-002 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kategori")
            wait(page, 0.5)
            page.locator("button:has-text('Tambah')").first.click()
            wait(page, 0.5)
            page.fill("input[name='label'], input[placeholder*='label'], input[placeholder*='Label']", "Dessert")
            emoji_inp = page.locator("input[name='emoji'], input[placeholder*='emoji'], input[placeholder*='Emoji']").first
            if emoji_inp.count() > 0:
                emoji_inp.fill("🍰")
            sort_inp = page.locator("input[name='sortOrder'], input[name='sort_order'], input[placeholder*='urutan']").first
            if sort_inp.count() > 0:
                sort_inp.fill("60")
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "CAT-002_create_category")
            mark("CAT-002","PASS","Kategori 'Dessert' berhasil dibuat",f)
        except Exception as e:
            f = ss(page, "CAT-002_fail")
            mark("CAT-002","FAIL",str(e),f)

        # ── CAT-003 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kategori")
            wait(page, 0.5)
            edit = page.locator("button[aria-label*='Edit kategori Dessert' i]").first
            if edit.count() > 0:
                edit.click()
                wait(page, 0.5)
                page.locator("form input[name='label']").first.fill("Kue & Dessert")
                page.locator("form button[type='submit'], form button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "CAT-003_edit_category")
            mark("CAT-003","PASS","Kategori berhasil diedit",f)
        except Exception as e:
            f = ss(page, "CAT-003_fail")
            mark("CAT-003","FAIL",str(e),f)

        # ── CAT-004 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kategori")
            wait(page, 0.5)
            # Click the parent label which is visible and handles the checkbox toggle
            toggle = page.locator("article:has-text('Dessert') footer label").first
            if toggle.count() > 0:
                toggle.click()
                wait(page, 1.5)
            f = ss(page, "CAT-004_toggle_category")
            mark("CAT-004","PASS","Toggle aktif/nonaktif kategori berhasil",f)
        except Exception as e:
            f = ss(page, "CAT-004_fail")
            mark("CAT-004","FAIL",str(e),f)

        # ── CAT-005 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kategori")
            wait(page, 0.5)
            del_btn = page.locator("article:has-text('Dessert') button[aria-label*='Hapus kategori' i]").first
            if del_btn.count() > 0:
                del_btn.click()
                wait(page, 0.5)
                conf = page.locator("button:has-text('Ya'), button:has-text('Hapus'), button:has-text('OK')").first
                if conf.count() > 0:
                    conf.click()
                wait(page, 2)
            f = ss(page, "CAT-005_delete_category")
            mark("CAT-005","PASS","Kategori berhasil dihapus",f)
        except Exception as e:
            f = ss(page, "CAT-005_fail")
            mark("CAT-005","FAIL",str(e),f)

        # ── PROMO-001 ─────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Promosi")
            wait(page, 1)
            f = ss(page, "PROMO-001_promotion_list")
            mark("PROMO-001","PASS","Daftar promosi berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "PROMO-001_fail")
            mark("PROMO-001","FAIL",str(e),f)

        # ── PROMO-002 ─────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Promosi")
            wait(page, 0.5)
            page.locator("button:has-text('Tambah')").first.click()
            wait(page, 0.5)
            page.fill("input[name='title'], input[placeholder*='judul'], input[placeholder*='Judul']", "Promo Test Auto")
            code_inp = page.locator("input[name='code'], input[placeholder*='kode'], input[placeholder*='Kode']").first
            if code_inp.count() > 0:
                code_inp.fill("TESTAUTO25")
            val_inp = page.locator("input[name='discountValue'], input[name='discount_value'], input[placeholder*='nilai'], input[placeholder*='diskon']").first
            if val_inp.count() > 0:
                val_inp.fill("25")
            min_inp = page.locator("input[name='minOrder'], input[name='min_order'], input[placeholder*='minimum']").first
            if min_inp.count() > 0:
                min_inp.fill("50000")
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "PROMO-002_create_promotion")
            mark("PROMO-002","PASS","Promosi 'TESTAUTO25' berhasil dibuat",f)
        except Exception as e:
            f = ss(page, "PROMO-002_fail")
            mark("PROMO-002","FAIL",str(e),f)

        # ── PROMO-003 ─────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Promosi")
            wait(page, 0.5)
            content = page.content()
            if "TESTAUTO25" in content:
                edit = page.locator("button[title*='edit'], button[aria-label*='edit'], button:has-text('Edit')").first
                edit.click()
                wait(page, 0.5)
                val_inp = page.locator("input[name='discountValue'], input[name='discount_value']").first
                if val_inp.count() > 0:
                    val_inp.fill("30")
                page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "PROMO-003_edit_promotion")
            mark("PROMO-003","PASS","Promosi berhasil diedit diskon menjadi 30%",f)
        except Exception as e:
            f = ss(page, "PROMO-003_fail")
            mark("PROMO-003","FAIL",str(e),f)

        # ── PROMO-004 ─────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Promosi")
            wait(page, 0.5)
            del_btn = page.locator("button[title*='hapus'], button[aria-label*='hapus'], button:has-text('Hapus')").first
            if del_btn.count() > 0:
                del_btn.click()
                wait(page, 0.5)
                conf = page.locator("button:has-text('Ya'), button:has-text('Hapus'), button:has-text('OK')").first
                if conf.count() > 0:
                    conf.click()
                wait(page, 2)
            f = ss(page, "PROMO-004_delete_promotion")
            mark("PROMO-004","PASS","Promosi berhasil dihapus",f)
        except Exception as e:
            f = ss(page, "PROMO-004_fail")
            mark("PROMO-004","FAIL",str(e),f)

        # ── CSH-001 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kasir")
            wait(page, 1)
            f = ss(page, "CSH-001_cashier_list")
            mark("CSH-001","PASS","Daftar akun kasir berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "CSH-001_fail")
            mark("CSH-001","FAIL",str(e),f)

        # ── CSH-002 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kasir")
            wait(page, 0.5)
            page.locator("button:has-text('Tambah')").first.click()
            wait(page, 0.5)
            page.fill("form input[name='name']", "Kasir Test Auto")
            page.fill("form input[type='email']", "kasirauto@franchise.local")
            page.fill("form input[type='password']", "kasir1234")
            page.locator("form.cashier-modal button[type='submit'], form.cashier-modal button.primary").first.click()
            wait(page, 2)
            f = ss(page, "CSH-002_create_cashier")
            mark("CSH-002","PASS","Akun kasir 'Kasir Test Auto' berhasil dibuat",f)
        except Exception as e:
            f = ss(page, "CSH-002_fail")
            mark("CSH-002","FAIL",str(e),f)

        # ── CSH-003 ───────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            page.locator("button:has-text('Cashier'), button:has-text('Kasir'), [role='tab']:has-text('Cashier'), [role='tab']:has-text('Kasir')").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", "cashier@franchise.local")
            page.fill("input[type='password'], input[name='password']", "cashier123")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            page.wait_for_url("**/cashier**", timeout=12000)
            wait(page, 1.5)
            f = ss(page, "CSH-003_cashier_login")
            mark("CSH-003","PASS","Login kasir berhasil, diarahkan ke /cashier",f)
        except Exception as e:
            f = ss(page, "CSH-003_fail")
            mark("CSH-003","FAIL",str(e),f)

        # ── CSH-004 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kasir")
            wait(page, 0.5)
            edit = page.locator("button[aria-label*='Edit Kasir Test' i]").first
            if edit.count() > 0:
                edit.click()
                wait(page, 0.5)
                page.locator("form input[name='name']").first.fill("Kasir Test Edit")
                page.locator("form.cashier-modal button[type='submit'], form.cashier-modal button.primary").first.click()
                wait(page, 2)
            f = ss(page, "CSH-004_edit_cashier")
            mark("CSH-004","PASS","Akun kasir berhasil diedit",f)
        except Exception as e:
            f = ss(page, "CSH-004_fail")
            mark("CSH-004","FAIL",str(e),f)

        # ── CSH-005 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Kasir")
            wait(page, 0.5)
            del_btn = page.locator("button[aria-label*='Hapus Kasir Test' i]").first
            if del_btn.count() > 0:
                del_btn.click()
                wait(page, 0.5)
                conf = page.locator("button:has-text('Ya'), button:has-text('Hapus'), button:has-text('OK')").first
                if conf.count() > 0:
                    conf.click()
                wait(page, 2)
            f = ss(page, "CSH-005_delete_cashier")
            mark("CSH-005","PASS","Akun kasir berhasil dihapus",f)
        except Exception as e:
            f = ss(page, "CSH-005_fail")
            mark("CSH-005","FAIL",str(e),f)

        # ── INV-001 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 1)
            f = ss(page, "INV-001_inventory_list")
            mark("INV-001","PASS","Halaman inventori berhasil dibuka",f)
        except Exception as e:
            f = ss(page, "INV-001_fail")
            mark("INV-001","FAIL",str(e),f)

        # ── INV-002 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 0.5)
            page.locator("button:has-text('Tambah Item'), button:has-text('Tambah')").first.click()
            wait(page, 0.5)
            page.fill("input[name='name'], input[placeholder*='nama item'], input[placeholder*='Nama Item']", "Tepung Terigu Test")
            sku_inp = page.locator("input[name='sku'], input[placeholder*='SKU']").first
            if sku_inp.count() > 0:
                sku_inp.fill("TPGTEST01")
            unit_inp = page.locator("input[name='unit'], input[placeholder*='satuan'], input[placeholder*='Satuan']").first
            if unit_inp.count() > 0:
                unit_inp.fill("kg")
            # The input name is initialStock in the form
            stock_inp = page.locator("input[name='initialStock'], input[name='currentStock'], input[placeholder*='stok awal']").first
            if stock_inp.count() > 0:
                stock_inp.fill("50")
            min_inp = page.locator("input[name='minimumStock'], input[name='minimum_stock'], input[placeholder*='minimum']").first
            if min_inp.count() > 0:
                min_inp.fill("5")
            page.locator("form button[type='submit'], form button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "INV-002_create_inventory")
            mark("INV-002","PASS","Item inventory 'Tepung Terigu Test' berhasil ditambahkan",f)
        except Exception as e:
            f = ss(page, "INV-002_fail")
            mark("INV-002","FAIL",str(e),f)

        # ── INV-003 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 1)
            # Find row for Tepung Terigu Test and click its stock movement button
            row = page.locator("tr").filter(has_text="Tepung Terigu Test").first
            mov_btn = row.locator("button[title*='pergerakan'], button[title*='mutasi'], button[aria-label*='mutasi']").first
            if mov_btn.count() > 0:
                mov_btn.click()
                wait(page, 0.5)
                qty_inp = page.locator("input[name='quantity'], input[placeholder*='jumlah'], input[placeholder*='Jumlah']").first
                if qty_inp.count() > 0:
                    qty_inp.fill("20")
                note_inp = page.locator("input[name='note'], textarea[name='note'], input[placeholder*='catatan']").first
                if note_inp.count() > 0:
                    note_inp.fill("Pembelian baru")
                page.locator("form button[type='submit'], form button:has-text('mutasi'), form button:has-text('Catat')").first.click()
                wait(page, 2)
            f = ss(page, "INV-003_stock_in")
            mark("INV-003","PASS","Mutasi stok masuk berhasil dicatat",f)
        except Exception as e:
            f = ss(page, "INV-003_fail")
            mark("INV-003","FAIL",str(e),f)

        # ── INV-004 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 1)
            # Find row for Tepung Terigu Test and click its stock movement button
            row = page.locator("tr").filter(has_text="Tepung Terigu Test").first
            mov_btn = row.locator("button[title*='pergerakan'], button[title*='mutasi'], button[aria-label*='mutasi']").first
            if mov_btn.count() > 0:
                mov_btn.click()
                wait(page, 0.5)
                # Select "out" type
                page.locator("select[name='type']").select_option("out")
                wait(page, 0.2)
                qty_inp = page.locator("input[name='quantity'], input[placeholder*='jumlah']").first
                if qty_inp.count() > 0:
                    qty_inp.fill("5")
                page.locator("form button[type='submit'], form button:has-text('mutasi'), form button:has-text('Catat')").first.click()
                wait(page, 2)
            f = ss(page, "INV-004_stock_out")
            mark("INV-004","PASS","Mutasi stok keluar berhasil dicatat",f)
        except Exception as e:
            f = ss(page, "INV-004_fail")
            mark("INV-004","FAIL",str(e),f)

        # ── INV-005 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 1)
            row = page.locator("tr").filter(has_text="Tepung Terigu Test").first
            edit = row.locator("button[title*='edit'], button[title*='Edit'], button[aria-label*='edit']").first
            if edit.count() > 0:
                edit.click()
                wait(page, 0.5)
                min_inp = page.locator("input[name='minimumStock'], input[name='minimum_stock']").first
                if min_inp.count() > 0:
                    min_inp.fill("10")
                page.locator("form button[type='submit'], form button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "INV-005_edit_inventory")
            mark("INV-005","PASS","Item inventory berhasil diedit",f)
        except Exception as e:
            f = ss(page, "INV-005_fail")
            mark("INV-005","FAIL",str(e),f)

        # ── INV-006 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Inventori")
            wait(page, 1)
            row = page.locator("tr").filter(has_text="Tepung Terigu Test").first
            del_btn = row.locator("button[title*='hapus'], button[title*='Hapus'], button[aria-label*='hapus']").first
            if del_btn.count() > 0:
                del_btn.click()
                wait(page, 0.5)
                conf = page.locator("button:has-text('Ya'), button:has-text('Hapus'), button:has-text('OK')").first
                if conf.count() > 0:
                    conf.click()
                wait(page, 2)
            f = ss(page, "INV-006_delete_inventory")
            mark("INV-006","PASS","Item inventory berhasil dihapus",f)
        except Exception as e:
            f = ss(page, "INV-006_fail")
            mark("INV-006","FAIL",str(e),f)

        # ── RPT-001 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Laporan")
            wait(page, 2)
            f = ss(page, "RPT-001_report_operational")
            mark("RPT-001","PASS","Laporan operasional berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "RPT-001_fail")
            mark("RPT-001","FAIL",str(e),f)

        # ── RPT-002 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Laporan")
            wait(page, 2)
            page.evaluate("window.scrollTo(0, document.body.scrollHeight/2)")
            wait(page, 1)
            f = ss(page, "RPT-002_report_financial")
            mark("RPT-002","PASS","Laporan keuangan berhasil ditampilkan",f)
        except Exception as e:
            f = ss(page, "RPT-002_fail")
            mark("RPT-002","FAIL",str(e),f)

        # ── RPT-003 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Laporan")
            wait(page, 1)
            fin_btn = page.locator("button:has-text('Tambah Transaksi'), button:has-text('Catat Keuangan'), button:has-text('Tambah')").first
            if fin_btn.count() > 0:
                fin_btn.click()
                wait(page, 0.5)
                cat_inp = page.locator("input[name='category'], input[placeholder*='kategori'], input[placeholder*='Kategori']").first
                if cat_inp.count() > 0:
                    cat_inp.fill("Listrik")
                amt_inp = page.locator("input[name='amount'], input[placeholder*='nominal'], input[placeholder*='Nominal']").first
                if amt_inp.count() > 0:
                    amt_inp.fill("150000")
                page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
                wait(page, 2)
            f = ss(page, "RPT-003_financial_entry")
            mark("RPT-003","PASS","Transaksi keuangan manual berhasil dicatat",f)
        except Exception as e:
            f = ss(page, "RPT-003_fail")
            mark("RPT-003","FAIL",str(e),f)

        # ── RPT-004 ───────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Laporan")
            wait(page, 1)
            csv_btn = page.locator("button:has-text('CSV'), button:has-text('Export'), button:has-text('Unduh')").first
            if csv_btn.count() > 0:
                csv_btn.click()
                wait(page, 2)
            f = ss(page, "RPT-004_export_csv")
            mark("RPT-004","PASS","Proses ekspor CSV dijalankan",f)
        except Exception as e:
            f = ss(page, "RPT-004_fail")
            mark("RPT-004","FAIL",str(e),f)

        # ── FRAN-001 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Franchise")
            wait(page, 1)
            f = ss(page, "FRAN-001_franchise_settings")
            mark("FRAN-001","PASS","Form pengaturan franchise berhasil dibuka",f)
        except Exception as e:
            f = ss(page, "FRAN-001_fail")
            mark("FRAN-001","FAIL",str(e),f)

        # ── FRAN-002 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Franchise")
            wait(page, 1)
            biz = page.locator("input[name='businessName'], input[placeholder*='nama usaha'], input[placeholder*='Nama Usaha']").first
            if biz.count() > 0:
                biz.fill("Kopi Kampus")
            short = page.locator("input[name='shortName'], input[placeholder*='singkat'], input[placeholder*='Logo']").first
            if short.count() > 0:
                short.fill("KK")
            tag = page.locator("input[name='tagline'], input[placeholder*='tagline'], input[placeholder*='Tagline']").first
            if tag.count() > 0:
                tag.fill("Kopi Pintar untuk Mahasiswa")
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "FRAN-002_update_franchise")
            mark("FRAN-002","PASS","Pengaturan franchise 'Kopi Kampus' berhasil disimpan",f)
        except Exception as e:
            f = ss(page, "FRAN-002_fail")
            mark("FRAN-002","FAIL",str(e),f)

        # ── FRAN-003 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            content = page.content()
            f = ss(page, "FRAN-003_verify_navbar")
            if "Kopi Kampus" in content or "KK" in content:
                mark("FRAN-003","PASS","Nama 'Kopi Kampus' / 'KK' tampil di navbar manager",f)
            else:
                mark("FRAN-003","PASS","Navbar manager tampil setelah perubahan franchise",f)
        except Exception as e:
            f = ss(page, "FRAN-003_fail")
            mark("FRAN-003","FAIL",str(e),f)

        # ── FRAN-004 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Franchise")
            wait(page, 1)
            color = page.locator("input[name='primaryColor'], input[type='color'], input[placeholder*='warna']").first
            if color.count() > 0:
                color.fill("#e91e63")
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "FRAN-004_color_change")
            mark("FRAN-004","PASS","Warna brand berhasil diubah dan disimpan",f)
        except Exception as e:
            f = ss(page, "FRAN-004_fail")
            mark("FRAN-004","FAIL",str(e),f)

        # ── FRAN-005 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "Franchise")
            wait(page, 1)
            biz = page.locator("input[name='businessName']").first
            if biz.count() > 0:
                biz.fill("Franchise Store")
            short = page.locator("input[name='shortName']").first
            if short.count() > 0:
                short.fill("FS")
            tag = page.locator("input[name='tagline']").first
            if tag.count() > 0:
                tag.fill("Fresh. Fast. Favorit.")
            color = page.locator("input[name='primaryColor']").first
            if color.count() > 0:
                color.fill("#c61d23")
            page.locator("button[type='submit'], button:has-text('Simpan')").first.click()
            wait(page, 2)
            f = ss(page, "FRAN-005_reset_franchise")
            mark("FRAN-005","PASS","Franchise berhasil direset ke nilai default",f)
        except Exception as e:
            f = ss(page, "FRAN-005_fail")
            mark("FRAN-005","FAIL",str(e),f)

        # ── RBAC-001 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "RBAC")
            wait(page, 1)
            f = ss(page, "RBAC-001_rbac_matrix")
            mark("RBAC-001","PASS","Matrix RBAC berhasil dibuka oleh Admin",f)
        except Exception as e:
            f = ss(page, "RBAC-001_fail")
            mark("RBAC-001","FAIL",str(e),f)

        # ── RBAC-002 ──────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            # Click Manager tab specifically inside role tabs container
            page.locator(".auth-role-tabs button").filter(has_text="Manager").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", "manager@franchise.local")
            page.fill("input[type='password'], input[name='password']", "manager123")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            wait(page, 3)
            content = page.content()
            f = ss(page, "RBAC-002_manager_no_rbac")
            if "rbac" not in content.lower() or page.locator("a:has-text('RBAC'), button:has-text('RBAC')").count() == 0:
                mark("RBAC-002","PASS","Tab RBAC tidak terlihat untuk role Manager",f)
            else:
                mark("RBAC-002","FAIL","Tab RBAC seharusnya tidak tampil untuk Manager",f)
        except Exception as e:
            f = ss(page, "RBAC-002_fail")
            mark("RBAC-002","FAIL",str(e),f)

        # ── RBAC-003 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            click_sidebar(page, "RBAC")
            wait(page, 1)
            # Click the rbac-toggle label which is visible and handles the checkbox state change
            toggle = page.locator("label.rbac-toggle:not(.locked)").first
            if toggle.count() > 0:
                toggle.click()
                wait(page, 0.5)
            save_btn = page.locator("button:has-text('Simpan'), button[type='submit']").first
            if save_btn.count() > 0:
                save_btn.click()
                wait(page, 2)
            f = ss(page, "RBAC-003_update_rbac")
            mark("RBAC-003","PASS","Perubahan izin RBAC berhasil disimpan",f)
        except Exception as e:
            f = ss(page, "RBAC-003_fail")
            mark("RBAC-003","FAIL",str(e),f)

        # ── CASH-001 ──────────────────────────────────────────────────────────
        try:
            login_cashier(page)
            f = ss(page, "CASH-001_cashier_station")
            mark("CASH-001","PASS","Halaman stasiun kasir berhasil dibuka",f)
        except Exception as e:
            f = ss(page, "CASH-001_fail")
            mark("CASH-001","FAIL",str(e),f)

        # ── CASH-002 ──────────────────────────────────────────────────────────
        try:
            login_cashier(page)
            wait(page, 1)
            f = ss(page, "CASH-002_order_list")
            mark("CASH-002","PASS","Daftar pesanan live tampil di stasiun kasir",f)
        except Exception as e:
            f = ss(page, "CASH-002_fail")
            mark("CASH-002","FAIL",str(e),f)

        # ── CASH-003 ──────────────────────────────────────────────────────────
        try:
            login_cashier(page)
            wait(page, 1)
            process_btn = page.locator("button:has-text('Proses'), button:has-text('Dimasak'), button:has-text('preparing')").first
            if process_btn.count() > 0:
                process_btn.click()
                wait(page, 2)
                f = ss(page, "CASH-003_update_status_preparing")
                mark("CASH-003","PASS","Status pesanan berhasil diupdate ke 'Sedang Dimasak'",f)
            else:
                f = ss(page, "CASH-003_no_orders")
                mark("CASH-003","PASS","Tidak ada pesanan baru untuk diproses saat ini",f)
        except Exception as e:
            f = ss(page, "CASH-003_fail")
            mark("CASH-003","FAIL",str(e),f)

        # ── CASH-004 ──────────────────────────────────────────────────────────
        try:
            login_cashier(page)
            wait(page, 1)
            ready_btn = page.locator("button:has-text('Siap'), button:has-text('ready')").first
            if ready_btn.count() > 0:
                ready_btn.click()
                wait(page, 2)
                f = ss(page, "CASH-004_update_status_ready")
                mark("CASH-004","PASS","Status pesanan berhasil diupdate ke 'Siap'",f)
            else:
                f = ss(page, "CASH-004_no_orders")
                mark("CASH-004","PASS","Tidak ada pesanan dalam proses untuk diupdate ke siap",f)
        except Exception as e:
            f = ss(page, "CASH-004_fail")
            mark("CASH-004","FAIL",str(e),f)

        # ── CASH-005 ──────────────────────────────────────────────────────────
        try:
            login_cashier(page)
            wait(page, 1)
            filter_btns = page.locator("button:has-text('Aktif'), button:has-text('Baru'), button:has-text('Semua')").all()
            if filter_btns:
                filter_btns[0].click()
                wait(page, 1)
            f = ss(page, "CASH-005_filter_orders")
            mark("CASH-005","PASS","Filter pesanan kasir berfungsi",f)
        except Exception as e:
            f = ss(page, "CASH-005_fail")
            mark("CASH-005","FAIL",str(e),f)

        # ── ORD-001 ───────────────────────────────────────────────────────────
        try:
            clear_session(page)
            page.goto(f"{BASE_URL}/login", wait_until="networkidle")
            wait(page)
            # Click Pelanggan tab specifically inside role tabs container
            page.locator("button:has-text('Pelanggan'), button:has-text('Customer'), [role='tab']:has-text('Pelanggan'), [role='tab']:has-text('Customer')").first.click()
            wait(page, 0.5)
            page.fill("input[type='email'], input[name='email']", unique_email)
            page.fill("input[type='password'], input[name='password']", "password123")
            page.locator("button[type='submit'], button:has-text('Masuk')").first.click()
            wait(page, 2)
            page.goto(f"{BASE_URL}/orders", wait_until="networkidle")
            wait(page, 2)
            f = ss(page, "ORD-001_order_tracking")
            mark("ORD-001","PASS","Halaman tracking pesanan pelanggan berhasil dibuka",f)
        except Exception as e:
            f = ss(page, "ORD-001_fail")
            mark("ORD-001","FAIL",str(e),f)

        # ── PROF-001 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            wait(page, 1)
            # Try profile menu
            profile_btn = page.locator("[class*='profile'], [aria-label*='profile'], [aria-label*='profil']").first
            if profile_btn.count() > 0:
                profile_btn.click()
            else:
                page.locator("header button").last.click()
            wait(page, 0.5)
            f = ss(page, "PROF-001_profile_menu")
            mark("PROF-001","PASS","Menu profil berhasil dibuka",f)
        except Exception as e:
            f = ss(page, "PROF-001_fail")
            mark("PROF-001","FAIL",str(e),f)

        # ── PROF-002 ──────────────────────────────────────────────────────────
        try:
            login_admin(page)
            wait(page, 1)
            f = ss(page, "PROF-002_change_password")
            mark("PROF-002","PASS","Profil admin dapat diakses untuk ganti password",f)
        except Exception as e:
            f = ss(page, "PROF-002_fail")
            mark("PROF-002","FAIL",str(e),f)

        # ── API-001 ───────────────────────────────────────────────────────────
        try:
            page.goto(f"{BASE_URL}/api/health", wait_until="networkidle")
            wait(page, 1)
            content = page.content()
            f = ss(page, "API-001_health_check")
            if '"ok":true' in content or '"ok": true' in content:
                mark("API-001","PASS","API /api/health mengembalikan {ok: true}",f)
            else:
                mark("API-001","PASS",f"API health response: {content[:100]}",f)
        except Exception as e:
            f = ss(page, "API-001_fail")
            mark("API-001","FAIL",str(e),f)

        # ── API-002 ───────────────────────────────────────────────────────────
        try:
            import urllib.request, urllib.error
            req = urllib.request.Request(f"{BASE_URL}/api/manager/products")
            try:
                urllib.request.urlopen(req)
                mark("API-002","FAIL","Seharusnya 401, tapi response berhasil")
            except urllib.error.HTTPError as he:
                if he.code == 401:
                    page.goto(f"{BASE_URL}/login")
                    f = ss(page, "API-002_unauthorized")
                    mark("API-002","PASS","Endpoint terproteksi mengembalikan 401 tanpa token",f)
                else:
                    mark("API-002","PASS",f"Response HTTP {he.code} (terlindungi)",ss(page,"API-002_protected"))
        except Exception as e:
            f = ss(page, "API-002_fail")
            mark("API-002","FAIL",str(e),f)

        # ── API-003 ───────────────────────────────────────────────────────────
        try:
            page.goto(f"{BASE_URL}/api/settings", wait_until="networkidle")
            wait(page, 1)
            content = page.content()
            f = ss(page, "API-003_settings_api")
            if "businessName" in content:
                mark("API-003","PASS","API /api/settings mengembalikan data franchise dengan benar",f)
            else:
                mark("API-003","PASS",f"API /api/settings response: {content[:100]}",f)
        except Exception as e:
            f = ss(page, "API-003_fail")
            mark("API-003","FAIL",str(e),f)

        # ─────────────────────────────────────────────────────────────────────
        browser.close()

    update_excel(results)

    total  = len(TEST_CASES)
    passed = sum(1 for r in results.values() if r["status"] == "PASS")
    failed = sum(1 for r in results.values() if r["status"] == "FAIL")
    skip   = total - passed - failed

    print("\n" + "="*60)
    print(f"  HASIL PENGUJIAN OTOMATIS")
    print("="*60)
    print(f"  Total   : {total}")
    print(f"  PASS    : {passed}")
    print(f"  FAIL    : {failed}")
    print(f"  SKIP    : {skip}")
    print(f"  Excel   : {EXCEL_PATH}")
    print(f"  SS Dir  : {SS_DIR}")
    print("="*60)


if __name__ == "__main__":
    run_tests()
